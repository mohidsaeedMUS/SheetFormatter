import openpyxl
from openpyxl import Workbook

wb = openpyxl.load_workbook("Test_template.xlsx")
#ws=wb.active

for sheet in wb.sheetnames:
    print("Sheet: " + sheet)
    ws = wb[sheet]
    column_string=input('Enter Column Letter with First Name: ')
    iter_column=openpyxl.utils.cell.column_index_from_string(column_string) 
    column_string_last=input('Enter Column Letter with Last Name: ')
    iter_column_last=openpyxl.utils.cell.column_index_from_string(column_string_last)
    i = 1
    while i <= ws.max_row:
        if ws.cell(row=i, column=iter_column).value is None:
            #print(ws.cell(row=i, column=iter_column))
            ws.delete_rows(i, 1)
        else:
            i +=1
    j = 1
    while j <= ws.max_row:
        if ws.cell(row=j, column=iter_column_last).value is None:
            #print(ws.cell(row=i, column=iter_column_last))
            ws.delete_rows(j, 1)
        else:
            j +=1

wb.save("Test_complete.xlsx")