import openpyxl,pandas
wb = openpyxl.load_workbook("Test_template.xlsx")
max_rows_dict={}
max_columns_dict={}
sheet_num=0
count=0
for sheet in wb.sheetnames:
    data = pandas.read_excel("Test_template.xlsx",sheet_name=sheet)
    max_rows_dict[sheet_num]=len(data)+1
    max_columns_dict[sheet_num]=len(data.columns)
    sheet_num+=1
def access_val(num,dict_1,dict_2):
    return dict_1[num],dict_2[num]
for sheet in wb:
    max_row,max_col=access_val(count,max_rows_dict,max_columns_dict)
    count+=1
    column_string=input("Enter Column Letter with Email (A or B or C or leave blank to skip editing):").upper()
    if len(column_string)>0:
        first_time_blank=False
        for cell in sheet[column_string][1:max_row]:
            if cell.value is None:
                first_time_blank=True
        if first_time_blank==True:
            new_sh=wb.create_sheet('Linkedin Only')
            for i in range (1, max_row +1):
                for j in range (1, max_col + 1):
                    c = sheet.cell(row = i, column = j)
                    new_sh.cell(row = i, column = j).value = c.value
        # if two_sheets==True:
            for cell in new_sh[column_string][1:]:
                if cell.value is not None:
                    new_sh.delete_rows(cell.row)
            for cell in sheet[column_string][1:]:
                if cell.value is None:
                    sheet.delete_rows(cell.row)
wb.save("Test_1.xlsx")




#add condition to delete empty sheets at end