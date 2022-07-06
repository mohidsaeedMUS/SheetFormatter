import sys
import os
import pandas as pd
import openpyxl
print('Lead Sheet Formatter. Ctrl/Cmd + C to exit.')
file_path = input("Enter file path of Excel sheet: ")
assert os.path.exists(file_path), "File not found at " + str(file_path)
#connect sheets by a first name/ last name/ company name/ email and automatically use that as subset for duplicate?
#df = pd.concat(pd.read_excel('/Users/amark/Documents/GitHub/SheetFormatter/Final Test/Test_complete copy.xlsx', sheet_name=None),ignore_index=True)
df = pd.concat(pd.read_excel(str(file_path), sheet_name=None),ignore_index=True)
column_string=input('Enter Column Letter with First Name (or leave blank to skip): ').upper()
column_string_last=input('Enter Column Letter with Last Name (or leave blank to skip): ').upper()
i=1
column_names={}
for col in df.columns:
    column_names[i]=col
    i+=1
iter_column=openpyxl.utils.cell.column_index_from_string(column_string) 
iter_column_last=openpyxl.utils.cell.column_index_from_string(column_string_last) 
try:
    df.drop_duplicates(subset=[column_names[iter_column],column_names[iter_column_last]],keep='first',inplace=True)
except:
    df.drop_duplicates(keep='first',inplace=True)
max_rows,max_columns=len(df)+1,len(df.columns)
#df.to_excel('/Users/amark/Documents/GitHub/SheetFormatter/Final Test/Test_complete copy.xlsx',index=False)
df.to_excel(str(file_path),index=False)
#wb = openpyxl.load_workbook("/Users/amark/Documents/GitHub/SheetFormatter/Final Test/Test_complete copy.xlsx")
wb = openpyxl.load_workbook(str(file_path))
ws=wb.active
column_string_comp=input('Enter Column Letter with Company Name: (or leave blank to skip editing) ').upper()
#are exceptions truly exceptional?
dict_of_domains={0:'.co',1:'.io',2:'.net',3:'LLC',4:'.com',5:'.ai',6:'.org', 7:' , Inc.'}
try:
    for cell in ws[column_string_comp][1:max_rows]:
        z=0
        while z < 7:   
            if dict_of_domains[z] in cell.value:
                cell.value = cell.value.replace(dict_of_domains[z], "")
            z+=1
#is this ok?
except Exception:
    print('Invalid input entered or skipped deleting domains')
#is it exceptional or nor?
try:
    for cell in ws[column_string][1:max_rows]:
        if cell.value is None:
            ws.delete_rows(cell.row)
except Exception:
    print('Invalid input entered or skipped deleting rows with no first name. ')
    pass
try:
    for cell in ws[column_string_last][1:max_rows]:
        if cell.value is None:
            ws.delete_rows(cell.row)
except Exception:
    print('Invalid input entered or skipped deleting rows with no last name. ')
    pass
#is it exceptional?
column_string_email=input("Enter Column Letter with Email (or leave blank to skip editing): ").upper()
try:
    first_time_blank=False
    two_sheets=False
    for cell in ws[column_string_email][1:max_rows]:
        if cell.value is None:
            first_time_blank=True
            two_sheets=True
    if first_time_blank==True:
        ws_1=wb.create_sheet('Linkedin Only')
        for i in range (1, max_rows +1):
            for j in range (1, max_columns + 1):
                c = ws.cell(row = i, column = j)
                ws_1.cell(row = i, column = j).value = c.value
    if two_sheets==True:
        for cell in ws_1[column_string_email][1:]:
            if cell.value is not None:
                ws_1.delete_rows(cell.row)
        for cell in ws[column_string_email][1:]:
            if cell.value is None:
                ws.delete_rows(cell.row)
except Exception:
    print('Invalid input entered or skipped moving rows with no emails to another sheet.')
    pass
#wb.save("/Users/amark/Documents/GitHub/SheetFormatter/Final Test/Test_complete copy_result.xlsx")
completed_file_path = str(file_path).replace(".xlsx", " FINISHED.xlsx")
wb.save(str(completed_file_path))