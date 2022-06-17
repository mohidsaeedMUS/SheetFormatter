# import time
# start_time = time.perf_counter ()
import openpyxl
wb = openpyxl.load_workbook("Test.xlsx")
ws=wb.active
mr,mc=ws.max_row,ws.max_column
first_time_blank=False
two_sheets=False
column_string=input("Enter Column Letter with Email (A or B or C or leave blank to skip editing):").upper()
if len(column_string)>0:
    for cell in ws[column_string][1:]:
        if cell.value is None:
            first_time_blank=True
            two_sheets=True
    if first_time_blank==True:
        ws_1=wb.create_sheet('Linkedin Only')
        for i in range (1, mr +1):
            for j in range (1, mc + 1):
                c = ws.cell(row = i, column = j)
                ws_1.cell(row = i, column = j).value = c.value
    if two_sheets==True:
        for cell in ws_1[column_string][1:]:
            if cell.value is not None:
                ws_1.delete_rows(cell.row)
        for cell in ws[column_string][1:]:
            if cell.value is None:
                ws.delete_rows(cell.row)
    wb.save("Test.xlsx")
#need to be able to move on to next if left blank
# column=openpyxl.utils.cell.column_index_from_string(column_string) 
# for col in sh_new.iter_cols(min_row = 2, max_row = ws.max_row, min_col = column , max_col = column):
#     for cell in col:
#         rownum+=1
#         if cell.internal_value == "Verified":
# for row in sh_new.rows:
#     for cell in row:
#         if cell.value =="Verified":
#             sh_new.delete_rows(cell.row, 1)
else:
    wb.save("Test.xlsx")
# end_time = time.perf_counter ()
# print(end_time - start_time, "seconds")
#my worry- if email not empty but still want us to move it to linkedin only wont work 
#do we arrange each of these capabilites in functions?


    

