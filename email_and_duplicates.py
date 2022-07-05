import pandas as pd
import openpyxl
user_inp=input('Which headers should be considered as duplicate values? Example Answer: First Name/Last Name/Email (separate with /) ')
subset=user_inp.split("/")
for words in subset:
    words.strip()
#potentially set sheet name
df = pd.concat(pd.read_excel('/Users/amark/Documents/GitHub/SheetFormatter/Different_Headers.xlsx', sheet_name=None),ignore_index=True)
df.drop_duplicates(subset=subset,keep='first',inplace=True)
max_rows=len(df)+1
max_columns=len(df.columns)
#if it works change to test_dups_multiple and see if it stays in the same sheet
df.to_excel('/Users/amark/Documents/GitHub/SheetFormatter/Test Series 3/Test_both_copy.xlsx',index=False)
wb = openpyxl.load_workbook("/Users/amark/Documents/GitHub/SheetFormatter/Different_Headers.xlsx")
ws=wb.active
first_time_blank=False
two_sheets=False
column_string=input("Enter Column Letter with Email (A or B or C or leave blank to skip editing):").upper()
#maybe change to enter column name?
if len(column_string)>0:
    for cell in ws[column_string][1:max_columns]:
        if cell.value is None:
            first_time_blank=True
            two_sheets=True
    if first_time_blank==True:
        ws_1=wb.create_sheet('Linkedin Only')
        for i in range (1, max_rows +1):
            for j in range (1, max_columns + 1):
                c = ws.cell(row = i, column = j)
                ws_1.cell(row = i, column = j).value = c.value
    if two_sheets==True:
        for cell in ws_1[column_string][1:]:
            if cell.value is not None:
                ws_1.delete_rows(cell.row)
        for cell in ws[column_string][1:]:
            if cell.value is None:
                ws.delete_rows(cell.row)
wb.save("/Users/amark/Documents/GitHub/SheetFormatter/Different_Headers.xlsx")


