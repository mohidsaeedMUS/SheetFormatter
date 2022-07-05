import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill

wb = openpyxl.load_workbook("Test_template.xlsx")
ws=wb.active

redFill = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')

column_string=input('Enter Column Letter with Company Name: ')
iter_column=openpyxl.utils.cell.column_index_from_string(column_string) 

i = 1
while i <= ws.max_row:
    if len(ws.cell(row=i, column=iter_column).value) > 17:
        ws.cell(row=i, column=iter_column).fill = redFill
        i+=1
    else:
        i+=1

wb.save("Test_complete.xlsx")