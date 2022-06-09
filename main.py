import openpyxl
  
# load excel with its path
wrkbk = openpyxl.load_workbook("Test.xlsx")  
sh = wrkbk.active
sh_new = wrkbk.create_sheet(title = "Linkedin Only")
  
# iterate through excel and display data
# for col in sh.iter_cols(min_row = 2, max_row = 6, min_col = 5, max_col = 5):
#    for cell in col:
#        print(cell.value, end=" ")
#    print()

for col in sh.iter_cols(min_row = 2, max_row = 6, min_col = 5, max_col = 5):
    for cell in col:
        if cell.value.casefold() == "unavailable":
            print(True)
            break
    print()
wrkbk.save("Test.xlsx")
#test