import pandas as pd
import openpyxl
user_inp=input('Which headers should be considered as duplicate values? Example Answer: First Name/Last Name/Email (separate with /) ')
subset=user_inp.split("/")
#fix this (strip)
for words in subset:
    words.strip()
#potentially set sheet name
df = pd.concat(pd.read_excel('/Users/amark/Documents/GitHub/SheetFormatter/Test_complete copy.xlsx', sheet_name=None),ignore_index=True)
df.drop_duplicates(subset=subset,keep='first',inplace=True)
max_rows,max_columns=len(df)+1,len(df.columns)
#if it works change to test_dups_multiple and see if it stays in the same sheet
df.to_excel('/Users/amark/Documents/GitHub/SheetFormatter/Test_complete copy.xlsx',index=False)
wb = openpyxl.load_workbook("/Users/amark/Documents/GitHub/SheetFormatter/Test_complete copy.xlsx")
ws=wb.active
# .co/.io/.net/ LLC / .com/ .ai/ .org/ , Inc.
dict_of_domains={0:'.co',1:'.io',2:'.net',3:'LLC',4:'.com',5:'.ai',6:'.org', 7:' , Inc.'}
column_string=input('Enter Column Letter with Company Name: ').upper()
for cell in ws[column_string][1:max_rows]:
    z=0
    while z < 7:   
        if dict_of_domains[z] in cell.value:
                cell.value = cell.value.replace(dict_of_domains[z], "")
        z+=1

for sheet in wb.sheetnames:
    print("Sheet: " + sheet)
    ws = wb[sheet]
    column_string=input('Enter Column Letter with First Name: ')
    iter_column=openpyxl.utils.cell.column_index_from_string(column_string) 
    column_string_last=input('Enter Column Letter with Last Name: ')
    iter_column_last=openpyxl.utils.cell.column_index_from_string(column_string_last)
    i = 1
    while i <= ws.max_row:
        if ws.cell(row=i, column=iter_column).value is None:
            #print(ws.cell(row=i, column=iter_column))
            ws.delete_rows(i, 1)
        else:
            i +=1
    j = 1
    while j <= ws.max_row:
        if ws.cell(row=j, column=iter_column_last).value is None:
            #print(ws.cell(row=i, column=iter_column_last))
            ws.delete_rows(j, 1)
        else:
            j +=1
first_time_blank=False
two_sheets=False
column_string=input("Enter Column Letter with Email (A or B or C or leave blank to skip editing):").upper()
# maybe change to enter column name?
if len(column_string)>0:
    for cell in ws[column_string][1:max_rows]:
        if cell.value is None:
            first_time_blank=True
            two_sheets=True
    if first_time_blank==True:
        ws_1=wb.create_sheet('Linkedin Only')
        for i in range (1, max_rows +1):
            for j in range (1, max_columns + 1):
                c = ws.cell(row = i, column = j)
                ws_1.cell(row = i, column = j).value = c.value
    if two_sheets==True:
        for cell in ws_1[column_string][1:]:
            if cell.value is not None:
                ws_1.delete_rows(cell.row)
        for cell in ws[column_string][1:]:
            if cell.value is None:
                ws.delete_rows(cell.row)
wb.save("/Users/amark/Documents/GitHub/SheetFormatter/end.xlsx")