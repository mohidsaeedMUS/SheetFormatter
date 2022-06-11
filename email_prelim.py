import openpyxl
wb = openpyxl.load_workbook("Test_in_vscode.xlsx")
ws=wb.active
# column_string=input('Enter Column Letter with Email:')
#need to be able to move on to next if left blank
# column=openpyxl.utils.cell.column_index_from_string(column_string)
# count=0
# start=1
#grab a row and input it into other sheet then delete it 
# for row in ws.iter_rows(min_col=column,max_col=column,min_row=2,max_row=ws.max_row):
    # for cell in row:
    #     start+=1
    #     if cell.internal_value=='Unavailable':
    #         count+=1
    #         if count==1:
    #             ws1=wb.create_sheet('Linkedin Only')
    #             ws1.append(["First Name","Last Name","Company Name","Email","Email Status","First Phone","Employees","Industry","Person", "Linkedin","Company City","Company State","Company Count"])
    #             ws.delete_rows(start,1)
mr=ws.max_row
mc=ws.max_column
column_string=input('Enter Column Letter with Email:')
#need to be able to move on to next if left blank
column=openpyxl.utils.cell.column_index_from_string(column_string) 
count=0
row=1
# grab a row if meets condition and input it into other sheet then delete it from current sheet
for col in ws.iter_cols(min_row = 2, max_row = ws.max_row, min_col = column , max_col = column):
    for cell in col:
        row+=1
        if cell.internal_value == "Unavailable":
            count+=1
            if count==1:
                sh_new = wb.create_sheet(title = "Linkedin Only")
                sh_new.append(["First Name","Last Name","Company Name","Email","Email Status","First Phone","Employees","Industry","Person", "Linkedin","Company City","Company State","Company Count"])
                for i in range (row, row+1):
                    for j in range (1, mc + 1):
                        c = ws.cell(row = i, column = j)
                        sh_new.cell(row = i, column = j).value = c.value
            else:
                for i in range (row, row+1):
                    for j in range (1, mc + 1):
                        c = ws.cell(row = i, column = j)
                        sh_new.cell(row = i, column = j).value = c.value
                #grab row data
                #add to new sheet
                #delete from ws
wb.save('Test_in_vscode.xlsx')


 


